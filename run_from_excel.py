import os
import re
import openpyxl
from monsterStatBlock import MonsterStatBlock
from alterStatistics import alterStatistics

_RANGE_RE = re.compile(r"^\d+-\d+$")

BASE_DIR = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(BASE_DIR, "alter_statistics_inputs.xlsx")
INPUT_JSON = os.path.join(BASE_DIR, "guid_mapper_master.json")
OUTPUT_JSON = os.path.join(BASE_DIR, "guid_mapper_master_alteredStats.json")

ALTER_SHEET = "alter_stats"
CLEAR_SHEET = "clear_stats"

STATUS_READY = "ready"
STATUS_PROCESSED = "processed"
STATUS_NOT_READY = "not ready"

# Both sheets are read by column HEADER NAME from row 1, so column order doesn't matter.
#
# alter_stats recognised headers:
#   #, status, monster_archetype_match, class_archetype_match, subtype_match,
#   act_match, location_match, type_match, handle_match, full_guid_match,
#   corpse_boolean, passives_per, passives_to_add, spells_per, spells_to_add,
#   health_override, apply_base_armor_class (alias: apply_base_ac), ac_boost,
#   profiles_to_add
#
# clear_stats recognised headers:
#   #, status, monster_archetype_match, class_archetype_match, subtype_match,
#   act_match, location_match, type_match, handle_match, full_guid_match,
#   corpse_boolean, clear_passives, clear_spells, clear_health

CLEAR_HEADERS = [
    "#",
    "status",
    "monster_archetype_match",
    "clear_passives",
    "clear_spells",
    "clear_health",
    "class_archetype_match",
    "subtype_match",
    "act_match",
    "corpse_boolean",
    "handle_match",
    "full_guid_match",
    "location_match",
    "type_match",
]
CLEAR_PLACEHOLDER_ROWS = 8


def _str(value):
    return "" if value is None else str(value).strip()


def _int(value, default=0):
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _health_value(value):
    """Return value as-is if it's a 'N-M' range string, otherwise coerce to int."""
    if value is None:
        return 0
    s = str(value).strip()
    if _RANGE_RE.match(s):
        return s
    try:
        return int(s)
    except (ValueError, TypeError):
        return 0


def _bool_or_none(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no"):
        return False
    return None


def _list_cell(value):
    """Newline-delimited cell → list of strings."""
    if not value:
        return []
    return [item.strip() for item in str(value).split("\n") if item.strip()]


def _read_row(ws, row_idx):
    return [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]


def _build_header_map(ws):
    """Return {header_name: 0-based_index} from row 1 of the sheet. First occurrence wins."""
    result = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is not None:
            name = str(h).strip()
            if name not in result:
                result[name] = c - 1
    return result


def _process_alter_sheet(ws, blocks):
    hdr = _build_header_map(ws)

    def col(name, *aliases):
        """Look up a value from vals by header name (with optional fallback aliases)."""
        for n in (name,) + aliases:
            idx = hdr.get(n)
            if idx is not None and idx < len(vals):
                return vals[idx]
        return None

    status_col = hdr.get("status", 1) + 1  # 1-based for ws.cell

    for row_idx in range(2, ws.max_row + 1):
        vals = _read_row(ws, row_idx)
        status_cell = ws.cell(row=row_idx, column=status_col)

        has_content = (
            bool(col("passives_to_add"))
            or bool(col("spells_to_add"))
            or bool(col("health_override"))
            or bool(col("profiles_to_add"))
        )

        if not has_content:
            status_cell.value = STATUS_NOT_READY
            continue

        if col("status") != STATUS_READY:
            continue

        blocks = alterStatistics.alter_stats_by_field_combo(
            blocks,
            handle_match_phrase=_str(col("handle_match")),
            full_guid_match_phrase=_str(col("full_guid_match")),
            act_match_phrase=_str(col("act_match")),
            location_match_phrase=_str(col("location_match")),
            type_match_phrase=_str(col("type_match")),
            subtype_match_phrase=_str(col("subtype_match")),
            class_archetype_match_phrase=_str(col("class_archetype_match")),
            monster_archetype_match_phrase=_str(col("monster_archetype_match")),
            corpse_boolean=_bool_or_none(col("corpse_boolean")),
            passives_per_block=_int(col("passives_per")),
            passives_to_add=_list_cell(col("passives_to_add")),
            spells_per_block=_int(col("spells_per")),
            spells_to_add=_list_cell(col("spells_to_add")),
            health_override=_health_value(col("health_override")),
            ac_boost=_int(col("ac_boost")),
            apply_base_armor_class=_int(col("apply_base_armor_class", "apply_base_ac")),
            profiles_to_add=_list_cell(col("profiles_to_add")),
        )

        status_cell.value = STATUS_PROCESSED
        print(f"  alter_stats row #{col('#')}: applied")

    return blocks


def _process_clear_sheet(ws, blocks):
    hdr = _build_header_map(ws)

    def col(name, *aliases):
        for n in (name,) + aliases:
            idx = hdr.get(n)
            if idx is not None and idx < len(vals):
                return vals[idx]
        return None

    status_col = hdr.get("status", 1) + 1

    for row_idx in range(2, ws.max_row + 1):
        vals = _read_row(ws, row_idx)
        status_cell = ws.cell(row=row_idx, column=status_col)

        clear_passives = bool(col("clear_passives"))
        clear_spells = bool(col("clear_spells"))
        clear_health = bool(col("clear_health"))
        has_content = clear_passives or clear_spells or clear_health

        if not has_content:
            status_cell.value = STATUS_NOT_READY
            continue

        if col("status") != STATUS_READY:
            continue

        blocks = alterStatistics.clear_stats_by_field_combo(
            blocks,
            handle_match_phrase=_str(col("handle_match")),
            full_guid_match_phrase=_str(col("full_guid_match")),
            act_match_phrase=_str(col("act_match")),
            location_match_phrase=_str(col("location_match")),
            type_match_phrase=_str(col("type_match")),
            subtype_match_phrase=_str(col("subtype_match")),
            class_archetype_match_phrase=_str(col("class_archetype_match")),
            monster_archetype_match_phrase=_str(col("monster_archetype_match")),
            corpse_boolean=_bool_or_none(col("corpse_boolean")),
            clear_passives=clear_passives,
            clear_spells=clear_spells,
            clear_health=clear_health,
        )

        status_cell.value = STATUS_PROCESSED
        print(f"  clear_stats row #{col('#')}: applied")

    return blocks


def _ensure_clear_sheet(wb):
    if CLEAR_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(CLEAR_SHEET)
    ws.append(CLEAR_HEADERS)
    for i in range(1, CLEAR_PLACEHOLDER_ROWS + 1):
        ws.append([i, STATUS_NOT_READY] + [None] * (len(CLEAR_HEADERS) - 2))


def main():
    blocks = MonsterStatBlock.load_from_json_file(INPUT_JSON)
    wb = openpyxl.load_workbook(EXCEL_PATH)

    _ensure_clear_sheet(wb)

    print("Processing clear_stats...")
    blocks = _process_clear_sheet(wb[CLEAR_SHEET], blocks)

    print("Processing alter_stats...")
    blocks = _process_alter_sheet(wb[ALTER_SHEET], blocks)

    print("Reconciling profiles...")
    blocks = alterStatistics.reconcile_profiles(blocks)

    wb.save(EXCEL_PATH)
    print(f"Excel updated: {EXCEL_PATH}")

    MonsterStatBlock.save_to_json_file(blocks, OUTPUT_JSON)
    print(f"JSON saved: {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
